#!/usr/bin/env python3

# SG - Medium site router installs using Cisco IOS-XE routers running SDWAN code
# This code will be customer specific as each customer will have different template variables
# ---import tracker sheet and build template csv for import into vManage to cutdown on manual work required to deploy routers---
# To adapt this code there are two main sections that require updating:
# Section 1 is the definition of vmanage_dict - each dictionary key maps to a column header which is a variable in a template
# Section 2 is the main loop that reads the tracker sheet, manipulates the data and then writes it into the dictionary
# Section 3 performs postcode lookups to obtain GPS coords and gathers a list of routes required for DNAC - the vmanage-import.csv file is written


# openpxyl is a library for handing MS Excel files

import openpyxl

# pandas is used for working with csv files

import pandas as pd

# requests allows API calls - used to correct the UK Postcodes which have no space

import requests

# allows whois lookup for public subnets ; static routes are required on DNAC

from ipwhois import IPWhois

# some standard libraries

import pprint
import ipaddress
import sys
import re
import math
import pickle

def postcode_api(postcode_apilist):

	# Function for passing a list of postcodes to an external site for lookup
	# Correct the postcode format (missing space) and return long + lat values
	# Send API request, passing in postcodes as a list
	# NOTE Maximum 100 postcodes
	
	postcode_uri = 'https://api.postcodes.io/postcodes'
	
	# Raise an exception requests.HTTPException error is response is anything other than 200 (OK)
	    
	json={"postcodes": postcode_apilist}
	postcode_lookup = ''
	try:
		postcode_lookup = requests.post(
			postcode_uri, 
			json={"postcodes": postcode_apilist}
		)
		postcode_lookup.raise_for_status()
	except requests.exceptions.ConnectionError:
		print(f'\nConnection error connecting to {postcode_uri}\nvManage import sheet has not been updated\n')
		sys.exit()
	except requests.HTTPError as error:
		print(f'\nHTTP Error:\n{error}')
		sys.exit()

	return(postcode_lookup)

# Open the tracker sheet

tracker_wb_obj = openpyxl.load_workbook(
#    '/home/nicko/Sanctuary -Medium Site - Router config R0.1.xlsx')
# above if using virtualbox, below is the WSL mount point
#    '/mnt/c/tftproot/Sanctuary -Medium Site - Router config R0.1.xlsx')
    '/mnt/c/Users/nick.oneill/Downloads/Sanctuary -Medium Site - Router config R0.1.xlsx')
tracker_sheet_obj = tracker_wb_obj.active

# initialise some variables

max_row = tracker_sheet_obj.max_row
public_29_list = []

# define dictionary for the vmanage-import.csv file with all dict keys mapping to csv columns, data will be apended to each key to form rows, one row per device

vmanage_dict = {'csv-deviceId': [],
                'csv-deviceIP': [],
                'csv-host-name': [],
                '/0/interface_and_tag/interface/if-name': [],
                '/0/interface_and_tag/interface/ip/address': [],
                '/0/vpn-instance/ip/route/vpn0_default_route/prefix': [],
                '/0/vpn-instance/ip/route/vpn0_default_route/next-hop/vpn0_next_hop/address': [],
                '//system/host-name': [],
                '//system/system-ip': [],
                '//system/site-id': [],
                '//system/gps-location/latitude': [],
                '//system/gps-location/longitude': [],
                #
                # This is the onboard sheet built
                # The next section of code will combine the full template details so the csv can be used for the onboarding template and the full template
                #
                '/500/Loopback0/interface/ip/address': [],
                '/100/Vlan5/interface/ip/address': [],
                '/100/Vlan10/interface/ip/address': [],
                '/100/Vlan218/interface/ip/address': [],
                '/0/interface_and_tag/interface/description': [],
                '/0/interface_and_tag/interface/shaping-rate': [],
                '/0/interface_and_tag/interface/bandwidth-downstream': [],
                '/500/Vlan3901/interface/ip/address': [],
                '/500/Vlan3901//dhcp-server/address-pool': [],
                '/500/Vlan3901//dhcp-server/options/default-gateway': [],
                '//switchport/interface/GigabitEthernet0/1/4/shutdown': [],
                '//switchport/interface/GigabitEthernet0/1/5/shutdown': []}


# main loop - loop through the tracker sheet and build rows for the vmanage-import.csv dictionary transforming some of the data

tracker_row = 2
postcode_list = []
print (f'{max_row} rows found ...\n')


while tracker_row <= max_row:
    # get the WAN IP and skip this row if not supplied (no circuit details)
    cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=8)
    if cell_obj.value is None:
        print(f'Row {tracker_row} has no circuit info - skipping')
        tracker_row = tracker_row + 1
        continue

    # get the serial no and check it's not a blank row
    cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=4)
    serial_no = cell_obj.value
    #
    if serial_no != None:
        # get routeable 29 network and add it to a list
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=11)
        routeable_29 = str(cell_obj.value)
        if '/' not in routeable_29: routeable_29 = routeable_29 + '/29'
        try:
            check_valid_ip_net = ipaddress.ip_network(routeable_29)
            public_29_list.append(routeable_29.split('/')[0])
        except:
            ValueError
        # get the router model number
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=3)
        device_type = cell_obj.value
        if device_type is None:
            print(f'Device type missing on row {tracker_row}')
            tracker_row = tracker_row + 1
            continue
        device_id = str(device_type) + '-' + str(serial_no)
        #print(f'Row :{tracker_row}  device {device_id}')
        # get the management ip/system ip
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=14)
        loopback_ip = str(cell_obj.value)
        if loopback_ip is None: 
            print(f'''>>> Missing loopback IP for {device_id}  Row {tracker_row} <<< Please fix and rerun''')
        if '/' not in loopback_ip: loopback_ip = loopback_ip + '/32'
        device_ip = loopback_ip.split('/')[0]
        # get the wan ip/prefix length
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=8)
        if '/' not in cell_obj.value: cell_obj.value = cell_obj.value + '/31' # if no prefix length assume a /31
        wan_ip = ipaddress.ip_interface(cell_obj.value)
        next_hop = (wan_ip.ip) - 1
        if str(wan_ip) == "31.119.4.50/31":
            # fix the BT f*** up! for this site - IP's wrong way around
            next_hop = (wan_ip.ip) + 1
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=5)
        mpls = False
        if cell_obj.value == 'MPLS':
            next_hop = (wan_ip.ip) + 1 # For MPLS circuits the PE has the odd numbered IP which is the next IP up so +1 instead of -1
            mpls = True
        # get VLAN tag for WAN port and combine with G0/0/0
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=10)
        wan_tag = str(cell_obj.value)
        wan_tag = wan_tag.lower()
        if wan_tag == '': wan_tag = 'none'
        if wan_tag == 'none':
            wan_if = 'GigabitEthernet0/0/0'
        else:
            wan_if = 'GigabitEthernet0/0/0' + '.' + str(wan_tag)
        # get router hostname
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=13)
        hostname = str(cell_obj.value)
        if " " in hostname:
            hostname = hostname.replace(" ","")
            print(f'Removed spaces from device: {serial_no} {hostname}')
        # if hostname is missing '-R1' add it
        if "-R1" not in hostname: hostname = hostname + "-R1"
        # extract the siteid from the hostname
        site_type = hostname.split('-')[1]
        site_ref = hostname.split('-')[2]
        site_id = site_type + site_ref
        # extract the postcode from the hostname
        postcode = hostname.split('-')[3]
        postcode_list.append(postcode)
        # write the info into a dictionary which will be written to a csv - this concludes the info required for the 'onboard' templates
        vmanage_dict['csv-deviceId'].append(device_id)
        vmanage_dict['csv-deviceIP'].append(device_ip)
        vmanage_dict['csv-host-name'].append(hostname)
        vmanage_dict['/0/vpn-instance/ip/route/vpn0_default_route/prefix'].append('0.0.0.0/0')
        vmanage_dict['/0/vpn-instance/ip/route/vpn0_default_route/next-hop/vpn0_next_hop/address'].append(next_hop)
        vmanage_dict['/0/interface_and_tag/interface/if-name'].append(wan_if)
        vmanage_dict['//system/host-name'].append(hostname)
        vmanage_dict['//system/system-ip'].append(device_ip)
        vmanage_dict['//system/site-id'].append(site_id)
        vmanage_dict['/0/interface_and_tag/interface/ip/address'].append(wan_ip)
        #
        # This is the onboard sheet built
        # The next section of code will combine the full template details so the import sheet can be used for both onboard and full templates
        #
        vmanage_dict['/500/Loopback0/interface/ip/address'].append(loopback_ip)
        # get vlan5 subnet and add +1 for the router IP
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=15)
        try:
            vlan_net = ipaddress.ip_network(cell_obj.value)
        except ValueError:
            print(f'Bad Loopback IP on row {tracker_row}:{cell_obj.value}')
            sys.exit()
        # vlan_net is a subnet, the router needs to be assigned the first usable IP address
        # vlan_net[1] is the first usable IP in the network range, but it will have the prefix length stripped off so we add it back on
        vlan5 = str(vlan_net[1]) + '/' + str(vlan_net.prefixlen)
        vmanage_dict['/100/Vlan5/interface/ip/address'].append(vlan5)
        # vget vlan10 subnet and add +1 for the router IP
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=16)
        vlan_net = ipaddress.ip_network(cell_obj.value)
        vlan10 = str(vlan_net[1]) + '/' + str(vlan_net.prefixlen)
        vmanage_dict['/100/Vlan10/interface/ip/address'].append(vlan10)
        # vget vlan218 subnet and add +1 for the router IP
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=17)
        vlan_net = ipaddress.ip_network(cell_obj.value)
        vlan218 = str(vlan_net[1]) + '/' + str(vlan_net.prefixlen)
        vmanage_dict['/100/Vlan218/interface/ip/address'].append(vlan218)
        # get circuit ref
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=6)
        wan_desc = cell_obj.value
        vmanage_dict['/0/interface_and_tag/interface/description'].append(wan_desc)
        # get up/down bandwidths - if only one bandwidth use for both - convert M to Kb
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=7)
        bandwidths = str(cell_obj.value)
        if '/' in bandwidths:
            downstream = bandwidths.split('/')[0]
            upstream = bandwidths.split('/')[1]
        elif '_' in bandwidths:
            downstream = bandwidths.split('_')[0]
            upstream = bandwidths.split('_')[1]            
        else:
            downstream = bandwidths
            upstream = bandwidths
        downstream = downstream.split('M')[0]
        upstream = upstream.split('M')[0]
        # remove all non numeric chars from downstream and upstream
        downstream = re.sub('[^0-9]','', downstream)
        upstream = re.sub('[^0-9]','', upstream)
        # modify downstream for EoFTTP circuits - these have been uplifted to 200Mb (originally ordered as 100Mb)
        if 'EoFTTP' in bandwidths:
            downstream = '200'
        downstream = downstream + '000'
        upstream = upstream + '000'
        vmanage_dict['/0/interface_and_tag/interface/shaping-rate'].append(upstream)
        vmanage_dict['/0/interface_and_tag/interface/bandwidth-downstream'].append(downstream)
        # use the public routable network as the provisioning network (Vlan3901) - this guarentees uniqueness and is not routable on the Internet so no security issue
        cell_obj = tracker_sheet_obj.cell(row=tracker_row, column=11)
        if cell_obj.value is None and mpls is True:
            cell_obj.value = '1.0.0.0/29' # /29 not present for MPLS
        if '/' not in cell_obj.value: cell_obj.value = cell_obj.value + '/29' # if no prefix length assume a /29
        vlan_net = ipaddress.ip_network(cell_obj.value)
        vlan3901 = str(vlan_net[1]) + '/' + str(vlan_net.prefixlen)
        vmanage_dict['/500/Vlan3901/interface/ip/address'].append(vlan3901)
        vmanage_dict['/500/Vlan3901//dhcp-server/address-pool'].append(str(vlan_net))
        vmanage_dict['/500/Vlan3901//dhcp-server/options/default-gateway'].append(vlan3901.split("/")[0])
        # Enable the switch provisoining ports by setting Shutdown state to FALSE
        vmanage_dict['//switchport/interface/GigabitEthernet0/1/4/shutdown'].append('FALSE')
        vmanage_dict['//switchport/interface/GigabitEthernet0/1/5/shutdown'].append('FALSE')
        # all done, update the row number and run the loop again until all rows have been processed
    else:
        print(f'Row {tracker_row} has no router serial number - skipping')
        tracker_row = tracker_row + 1
        continue
    tracker_row = tracker_row + 1

# Break the postcode list into chunks of 100 as the API has a max 100 limit

latlist = []
longlist = []

while len(postcode_list) > 100:
    postcode_max100 = postcode_list[0:100]
    postcode_result = postcode_api(postcode_max100)
    postcode_df = pd.json_normalize(postcode_result.json()['result'],sep='_')
    # update the csv dictionary with the lat and long values returned by the API
    latlist = latlist + (postcode_df['result_latitude'].to_list())
    longlist = longlist+ (postcode_df['result_longitude'].to_list())
    # remove the postcodes we have lookedup from the list
    postcode_list = postcode_list[100:]

if len(postcode_list) > 0:
    postcode_result = postcode_api(postcode_list)
    postcode_df = pd.json_normalize(postcode_result.json()['result'],sep='_')
    latlist = latlist + (postcode_df['result_latitude'].to_list())
    longlist = longlist+ (postcode_df['result_longitude'].to_list())
    
# update the csv dictionary with the lat and long values returned by the API
vmanage_dict['//system/gps-location/latitude'] = latlist
vmanage_dict['//system/gps-location/longitude'] = longlist



index = 0
for a in vmanage_dict['//system/gps-location/latitude']:
    if math.isnan(a): 
        print(f'''>>> Issue with hostname/postcode >>> {vmanage_dict['//system/host-name'][index]} <<< Please fix and rerun''')
    index = index+1

# uncomment the next line if you wish to view the dictionary
#pprint.pprint(vmanage_dict)

# create the dataframe from the dictionary we built
df = pd.DataFrame(vmanage_dict)

# write the dataframe to a csv ready for import into vManage
df.to_csv('/mnt/c/Users/nick.oneill/Downloads/vmanage-import.csv', index=False)
# change dir from ~ to use windows dowloads folder

# iterate through public /29 networks and do a whois lookup to discover the network block it was assigned from to build a list of routing entries for DNAC

net_block_list = []

print(f'Performing whois lookup. Please wait ...',end='',flush=True)

for net29 in public_29_list:
    print('.',end='',flush=True)
    obj = IPWhois(net29)
    lookup_result = obj.lookup_whois()
    #pprint.pprint(lookup_result)
    #net_block = lookup_result['nets'][1]['cidr']
    net_block = lookup_result['asn_cidr']
    net_block_list.append(net_block)

# report any additional routes required since the last run
# DNAC routes are stored in dnac_routes.txt

dnac_routes_on_file = []
# using the set function removes duplicate items from the list
net_block_list = list(set(net_block_list))

try:
    with open('dnac_routes.txt', "r") as f:
        dnac_routes_on_file = [line.rstrip('\n') for line in f]    
except:
    IOError

# using the set function removes duplicate items from the list
routes_since_last_run = list(set(net_block_list)-set(dnac_routes_on_file))

print(f'\n\nThe following routes were not detected the last time this was run and need adding to DNAC (enterprise interface):')
pprint.pprint(routes_since_last_run)

with open('dnac_routes.txt', 'w') as f:
    for item in net_block_list:
        f.write(item + '\n')


# all done
print('\nvmanage-import.csv has been created :)\n')
